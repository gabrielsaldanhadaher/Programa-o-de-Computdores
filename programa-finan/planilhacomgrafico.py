# !pip install pandas openpyxl yfinance
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
ativo = "PETR4.SA"
periodo = "6mo"
df = yf.download(ativo, period=periodo)
if df.empty:
  raise Exception("Ativo não encontrado ou sem dados")
if isinstance(df.columns, pd.MultiIndex):
  df.columns = df.columns.get_level_values(0)
df.reset_index(inplace=True)
aporte_mensal = 500
cotas = 0
carteira = []
for i in range(len(df)):
  preco = float(df["Close"].loc[i])
  if i% 21 == 0:
    cotas += aporte_mensal / preco
  carteira.append(cotas * preco)
df["Carteira"] = carteira
arquivo = "relatório_financeiro.xlsx"
df.to_excel(arquivo, index=False)
wb = load_workbook(arquivo)
ws = wb.active
max_linha = len(df) + 1
#mapear colunas automaticamente
colunas = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
col_data = colunas["Date"]
col_close = colunas["Close"]
col_carteira = colunas["Carteira"]

#gráfico

graf1 = LineChart()
graf1.title = "Preço ativo"
dados = Reference(ws, min_col=col_close, min_row=1, max_row=max_linha)
datas = Reference(ws, min_col=col_data, min_row=2, max_row=max_linha)
graf1.add_data(dados, titles_from_data = True)
graf1.set_categories(datas)
ws.add_chart(graf1, "H2")

#gráfico carteira
graf2 = LineChart()
graf2.title = "Evolução da Carteira"
dados2 = Reference(ws, min_col=col_carteira, min_row=1, max_row=max_linha)
graf2.add_data(dados2, titles_from_data=True)
graf2.set_categories(datas)
ws.add_chart(graf2, "H20")
wb.save(arquivo)
print("Excel com gráfico criado!")

#download

from google.colab import files
files.download(arquivo)