from openpyxl import Workbook, load_workbook

ARQUIVO = "financeiro.xlsx"

def criar_planilha():
  wb = Workbook()
  ws = wb.active
  ws.title = "Investimento"

  #cabeçalho

  ws.append(["Cliente", "Valor Inicial", "Taxa (%)", "Meses", "Valor Futuro"])

  #dados exemplo
  ws.append(["Ana", 1000, 1.5, 12, ""])
  ws.append(["João", 2000, 2.0, 10, ""])
  ws.append(["Maria", 1500, 1.2, 8, ""])
  wb.save(ARQUIVO)
  print("Planilha criada com sucesso!")

def calcular_financeiro():
  wb = load_workbook(ARQUIVO)
  ws = wb["Investimento"]
  for linha in ws.iter_rows(min_row=2):
    nome = linha[0].value
    VP = linha[1].value
    taxa = linha[2].value/100
    meses = linha[3].value
    #fómrula
    VF = VP * (1 + taxa)**meses
    linha[4].value = round(VF, 2)
  wb.save(ARQUIVO)
  print("Financeiro calculado com sucesso!")

  #exibir dados

def mostrar_planilha():
  wb = load_workbook(ARQUIVO)
  ws = wb["Investimento"]
  for linha in ws.iter_rows(values_only=True):
    print(linha)

criar_planilha()
calcular_financeiro()
mostrar_planilha()